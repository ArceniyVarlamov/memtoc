#!/usr/bin/env python3
"""Freeze the full v2 tool_wrong mapping.

Assembles the final per-episode distractor mapping from the round 1-3
accepted rows and the round-5 curation labels, applies the two human
exclusion layers (structural target-validity sweep + web-grounded gold
fact-check), and writes a single frozen artifact with provenance.

Owner decisions applied (annotator_a, 2026-07-16):
  D1: every malformed target-validity pair is excluded (23 episodes);
  D2: every fact-check-incorrect gold pair is excluded (no gold rewriting);
  D3: on authoring rows the frozen value is annotator_b's custom_value, per row
      (both annotators' values are kept in provenance);
  D4: pick-mode rows require exact annotator agreement (asserted, not
      adjudicated -- round 5 came back 18/18).

The script is pure assembly: it introduces no new values and refuses to
freeze on any inconsistency (missing labels, pick disagreement, count
mismatches, rubric violations).
"""

from __future__ import annotations

import argparse
import hashlib
import json
from pathlib import Path

from openpyxl import load_workbook

SCHEMA = "memtoc-frozen-mapping-v1"
EXPERIMENT = "memtoc-mapping-v2"
DECISION_DATE = "2026-07-16"
MAX_CANDIDATES = 6

EXPECTED = {
    "accepted_rounds_1_3": 254,
    "v5_rows": 50,
    "malformed_pairs": 8,
    "malformed_episodes": 23,
    "total_targets": 304,
    "frozen_episodes": 276,
}


def norm_text(value) -> str:
    return " ".join(str(value).split()).lower()


def sha256_file(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def load_json(path: Path):
    with path.open() as handle:
        return json.load(handle)


def read_sheet(path: Path, sheet: str) -> list[dict]:
    workbook = load_workbook(path, data_only=True)
    rows = list(workbook[sheet].iter_rows(values_only=True))
    header = [str(cell) for cell in rows[0]]
    return [dict(zip(header, row)) for row in rows[1:] if any(cell is not None for cell in row)]


def load_accepted_rounds(mapping_dir: Path) -> list[dict]:
    """Rounds 1-2 come from the cumulative partial mapping; round 3 is
    reconstructed from its bridge + final labels (it was never materialised
    as a file of its own).
    """
    partial = load_json(mapping_dir / "accepted_mapping_partial.json")
    accepted = []
    for row in partial["rows"]:
        accepted.append(
            {
                "episode_id": row["episode_id"],
                "instance_id": int(row["instance_id"]),
                "hop_idx": int(row["hop_idx"]),
                "question": row["question"],
                "gold_answer": row["gold_answer"],
                "old_tool_output": row["old_tool_output"],
                "final_tool_output": row["proposed_tool_output"],
                "resolution": "accepted_round_1_2",
                "provenance": {"blind_row_id": row["blind_row_id"]},
            }
        )

    bridge3 = {row["blind_row_id"]: row for row in load_json(mapping_dir / "replacement_v3_bridge_internal.json")["rows"]}
    labels3 = load_json(mapping_dir / "round3_labels_raw.json")["rows"]
    round3_yes = [row for row in labels3 if row["final_yes_no"] == "yes"]
    for label in round3_yes:
        row = bridge3[label["blind_row_id"]]
        accepted.append(
            {
                "episode_id": row["episode_id"],
                "instance_id": int(row["instance_id"]),
                "hop_idx": int(row["hop_idx"]),
                "question": row["question"],
                "gold_answer": row["gold_answer"],
                "old_tool_output": row["old_tool_output"],
                "final_tool_output": row["proposed_tool_output"],
                "resolution": "accepted_round_3",
                "provenance": {"blind_row_id": row["blind_row_id"]},
            }
        )

    assert len(accepted) == EXPECTED["accepted_rounds_1_3"], len(accepted)
    return accepted


def load_exclusions(mapping_dir: Path, factcheck_files: list[Path]) -> tuple[dict, list[dict]]:
    """episode_id -> exclusion record, plus the incorrect-gold detail rows."""
    adjudication = load_json(mapping_dir / "target_validity" / "target_validity_adjudication.json")
    verdicts = adjudication["final_verdicts"]
    validity_bridge = load_json(mapping_dir / "target_validity" / "target_validity_bridge_internal.json")["rows"]
    pair_by_key = {(norm_text(r["question"]), norm_text(r["gold_answer"])): r for r in validity_bridge}
    pair_by_id = {r["pair_id"]: r for r in validity_bridge}

    excluded: dict[str, dict] = {}
    malformed_pairs = [pid for pid, verdict in verdicts.items() if verdict == "malformed"]
    assert len(malformed_pairs) == EXPECTED["malformed_pairs"], malformed_pairs
    for pid in malformed_pairs:
        for episode_id in pair_by_id[pid]["episode_ids"]:
            excluded[episode_id] = {"pair_id": pid, "reason": "malformed_target"}
    assert len(excluded) == EXPECTED["malformed_episodes"], len(excluded)

    incorrect_rows = []
    for path in factcheck_files:
        for row in read_sheet(path, "gold_factcheck"):
            verdict = norm_text(row["verdict_fact"])
            assert verdict in {"correct", "incorrect"}, (path.name, row["pair_id"], verdict)
            if verdict != "incorrect":
                continue
            assert row["source_url"], (path.name, row["pair_id"], "incorrect verdict without source_url")
            pair = pair_by_id[row["pair_id"]]
            assert verdicts[row["pair_id"]] == "valid", row["pair_id"]
            incorrect_rows.append(
                {
                    "pair_id": row["pair_id"],
                    "question": pair["question"],
                    "gold_answer": pair["gold_answer"],
                    "corrected_answer": row["corrected_answer"],
                    "source_url": row["source_url"],
                    "annotator_file": path.name,
                    "episode_ids": pair["episode_ids"],
                }
            )
            for episode_id in pair["episode_ids"]:
                assert episode_id not in excluded, episode_id
                excluded[episode_id] = {"pair_id": row["pair_id"], "reason": "incorrect_gold"}

    return excluded, incorrect_rows


def load_v5_resolutions(mapping_dir: Path, annotator_a_xlsx: Path, annotator_b_xlsx: Path, excluded: dict) -> tuple[list[dict], list[str]]:
    bridge = {row["blind_row_id"]: row for row in load_json(mapping_dir / "curation_v5" / "curation_v5_bridge_internal.json")["rows"]}
    labels_a = {row["blind_row_id"]: row for row in read_sheet(annotator_a_xlsx, "QC8_curation")}
    labels_r = {row["blind_row_id"]: row for row in read_sheet(annotator_b_xlsx, "QC8_curation")}
    assert set(bridge) == set(labels_a) == set(labels_r)
    assert len(bridge) == EXPECTED["v5_rows"]

    def parse_pick(raw):
        text = norm_text(raw) if raw is not None else ""
        if text in {"", "none", "nan"}:
            return None
        return int(float(text))

    resolutions, skipped = [], []
    for blind_id, row in bridge.items():
        if row["episode_id"] in excluded:
            skipped.append(blind_id)
            continue
        label_a, label_r = labels_a[blind_id], labels_r[blind_id]
        pick_a, pick_r = parse_pick(label_a["pick"]), parse_pick(label_r["pick"])
        candidates = row["candidates"]

        base = {
            "episode_id": row["episode_id"],
            "instance_id": int(row["instance_id"]),
            "hop_idx": int(row["hop_idx"]),
            "question": row["question"],
            "gold_answer": row["gold_answer"],
            "old_tool_output": row["old_tool_output"],
        }
        if pick_a is not None or pick_r is not None:
            # pick mode: exact agreement is a freeze precondition (D4)
            assert pick_a == pick_r, (blind_id, pick_a, pick_r)
            assert 1 <= pick_a <= min(len(candidates), MAX_CANDIDATES), (blind_id, pick_a)
            chosen = candidates[pick_a - 1]
            base.update(
                final_tool_output=chosen["proposed_output"],
                resolution="v5_pick",
                provenance={
                    "blind_row_id": blind_id,
                    "pick": pick_a,
                    "source_instance_id": chosen["source_instance_id"],
                    "source_hop_idx": chosen["source_hop_idx"],
                    "source_tool_name": chosen["source_tool_name"],
                },
            )
        else:
            custom_a = str(label_a["custom_value"] or "").strip()
            custom_r = str(label_r["custom_value"] or "").strip()
            assert custom_a and custom_r, (blind_id, "authoring row without custom_value")
            for custom in (custom_a, custom_r):
                assert norm_text(custom) != norm_text(row["gold_answer"]), (blind_id, "custom repeats gold")
                assert all(norm_text(custom) != norm_text(c["proposed_output"]) for c in candidates), blind_id
            base.update(
                final_tool_output=custom_r,  # D3: authored values — annotator_b
                resolution="v5_author_annotator_b",
                provenance={
                    "blind_row_id": blind_id,
                    "custom_annotator_b": custom_r,
                    "custom_annotator_a": custom_a,
                },
            )
        resolutions.append(base)
    return resolutions, skipped


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--mapping-dir", type=Path, required=True)
    parser.add_argument("--out", type=Path, required=True)
    args = parser.parse_args()
    mapping_dir = args.mapping_dir

    factcheck_files = [
        mapping_dir / "gold_factcheck" / "checked" / "gold_factcheck_annotator_a_checked.xlsx",
        mapping_dir / "gold_factcheck" / "checked" / "gold_factcheck_annotator_b_checked.xlsx",
    ]
    v5_annotator_a = mapping_dir / "curation_v5" / "checked" / "QC8_curation_v5_annotator_a_completed.xlsx"
    v5_annotator_b = mapping_dir / "curation_v5" / "checked" / "QC8_curation_v5_annotator_b_completed.xlsx"

    accepted = load_accepted_rounds(mapping_dir)
    excluded, incorrect_rows = load_exclusions(mapping_dir, factcheck_files)
    v5_rows, v5_skipped = load_v5_resolutions(mapping_dir, v5_annotator_a, v5_annotator_b, excluded)

    kept_accepted = [row for row in accepted if row["episode_id"] not in excluded]
    frozen = kept_accepted + v5_rows

    episode_ids = [row["episode_id"] for row in frozen]
    assert len(episode_ids) == len(set(episode_ids)), "duplicate episodes in frozen mapping"
    assert len(frozen) + len(excluded) == EXPECTED["total_targets"], (len(frozen), len(excluded))
    assert len(frozen) == EXPECTED["frozen_episodes"], len(frozen)

    resolution_counts: dict[str, int] = {}
    for row in frozen:
        resolution_counts[row["resolution"]] = resolution_counts.get(row["resolution"], 0) + 1
    exclusion_counts: dict[str, int] = {}
    for record in excluded.values():
        exclusion_counts[record["reason"]] = exclusion_counts.get(record["reason"], 0) + 1

    artifact = {
        "schema_version": SCHEMA,
        "experiment": EXPERIMENT,
        "status": "frozen",
        "decision_date": DECISION_DATE,
        "decisions": {
            "D1_malformed_targets": "all 8 malformed pairs excluded (target-validity sweep, kappa 0.9298)",
            "D2_incorrect_golds": "all fact-check-incorrect pairs excluded; no gold rewriting",
            "D3_authoring_rows": "frozen value = annotator_b's custom_value per row; both values in provenance",
            "D4_pick_rows": "exact annotator agreement asserted (18/18 in round 5)",
            "owner": "annotator_a",
        },
        "inputs": {
            "gold_factcheck_sha256": {p.name: sha256_file(p) for p in factcheck_files},
            "curation_v5_sha256": {p.name: sha256_file(p) for p in (v5_annotator_a, v5_annotator_b)},
        },
        "counts": {
            "frozen_episodes": len(frozen),
            "excluded_episodes": len(excluded),
            "resolutions": resolution_counts,
            "exclusions": exclusion_counts,
            "v5_rows_skipped_as_excluded": len(v5_skipped),
        },
        "incorrect_golds": incorrect_rows,
        "excluded_episodes": [
            {"episode_id": episode_id, **record} for episode_id, record in sorted(excluded.items())
        ],
        "rows": sorted(frozen, key=lambda row: (row["instance_id"], row["hop_idx"])),
    }

    args.out.write_text(json.dumps(artifact, ensure_ascii=False, indent=1))
    summary = {key: artifact[key] for key in ("schema_version", "experiment", "status", "decision_date", "counts")}
    summary_path = args.out.with_name(args.out.stem + "_summary.json")
    summary_path.write_text(json.dumps(summary, ensure_ascii=False, indent=1))
    print(json.dumps(summary["counts"], ensure_ascii=False, indent=1))
    print(f"wrote {args.out} and {summary_path}")


if __name__ == "__main__":
    main()
