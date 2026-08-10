"""Package for MANUAL authoring of distractors: the rows the machine could not
handle (ambiguous gold, works, cross-cultural nobility) plus people with no
nationality in Wikidata. Each row comes with a short list of real Wikidata
candidates (same occupation/era) from which a human PICKS the best one or
writes their own. We select rather than invent from nothing.

Candidates are fetched by the same queries as the generator, out of a warm
cache, with almost no network traffic.

Run: python -m scripts.build_distractor_authoring
"""

from __future__ import annotations

import argparse
import json
from collections import Counter
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

from scripts.verify_gold_wikidata import Api, norm
from scripts.build_distractors_wikidata import (
    people_query, place_query, sparql, NEAR_ERA, FAME_CAP)

ROOT = Path(__file__).resolve().parent.parent

REASON_RU = {
    "gold_unresolved": "gold is ambiguous — pick by hand",
    "cross_culture_no_country_match": "cross-cultural nobility — needs a culturally appropriate one",
    "person_no_occupation_no_era": "no occupation/era in Wikidata",
    "kind_other_string_by_hand": "work",
    "kind_work_by_hand": "work (creative)",
    "nonentity_failed": "not typeable",
    "no_nationality": "person with no nationality in Wikidata",
}


def shortlist(api: Api, r: dict, k: int = 6) -> list[str]:
    """Up to k real Wikidata candidates to choose from; the less famous first."""
    prov = r.get("provenance", {}) or {}
    gold_norm = norm(r.get("gold_answer", ""))
    occ, sex, by = prov.get("occupations"), prov.get("sex"), prov.get("birth_year")
    types, country = prov.get("instance_of"), prov.get("country")
    if occ or sex:
        lo, hi = (by - NEAR_ERA, by + NEAR_ERA) if by else (None, None)
        rows = sparql(api, people_query(occ or [], sex, lo, hi, None))
    elif types:
        rows = sparql(api, place_query(types, country, same_country=True))
    else:
        return []
    out, seen = [], set()
    for b in rows:
        lbl = b.get("pLabel", {}).get("value", "")
        nl = norm(lbl)
        if not lbl or nl == gold_norm or nl in seen:
            continue
        if nl in gold_norm.split() or gold_norm in nl.split():
            continue
        sl = int(b["sl"]["value"]) if b.get("sl", {}).get("value") else 0
        out.append((sl, lbl))
        seen.add(nl)
    out.sort()  # the less famous (fewer language editions) first
    return [l for _, l in out[:k]]


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--dist", default=str(ROOT / "results" / "construction" / "distractors_v2_wikidata.json"))
    ap.add_argument("--cache", default=str(ROOT / "results" / "construction" / "wikidata_cache.json"))
    ap.add_argument("--out", default=str(ROOT / "results" / "construction" / "annotation" / "KCB_distractor_authoring.xlsx"))
    args = ap.parse_args()

    data = json.loads(Path(args.dist).read_text())["results"]
    api = Api(Path(args.cache))

    hand = []
    for r in data.values():
        if r.get("dropped"):
            continue
        if r.get("needs_human"):
            hand.append((r, r.get("reason", "?")))
        elif r.get("method") == "wikidata_person" and (r.get("provenance") or {}).get("country") is None:
            hand.append((r, "no_nationality"))
    print("to hand in total:", len(hand))
    print("by reason:", dict(Counter(rs for _, rs in hand).most_common()))

    # short lists (warm cache)
    for r, _ in hand:
        r["_short"] = shortlist(api, r)
    api.save()
    with_sl = sum(1 for r, _ in hand if r["_short"])
    print("with Wikidata candidates:", with_sl, "| free authoring:", len(hand) - with_sl)

    # sort: those with candidates first, then by family and id
    hand.sort(key=lambda x: (0 if x[0]["_short"] else 1, x[0]["family"],
                             x[0]["instance_id"], x[0]["hop_idx"]))

    wb = openpyxl.Workbook()
    head, wrap = Font(bold=True), Alignment(wrap_text=True, vertical="top")
    hi = PatternFill("solid", fgColor="FFF2CC")

    ws = wb.active
    ws.title = "Instructions"
    for i, (t, b) in enumerate([
        ("Manual authoring of distractors", True),
        ("", False),
        ("Every row needs ONE good near (the false tool value):", True),
        ("- of the same type/field/era as the correct answer,", False),
        ("- plausible (a model would reject it ONLY by knowing the correct answer),", False),
        ("- certainly wrong, and NOT a second admissible answer.", False),
        ("", False),
        ("The 'Wikidata candidates' column is a ready list to CHOOSE from (real entities",
         True),
        ("of the same occupation/era). Put a candidate number OR your own value in 'near'.", False),
        ("Where there are no candidates (ambiguous gold / a work), write your own.", False),
        ("", False),
        ("As before: two people independently, then we reconcile. 'checked' means the", False),
        ("second annotator marks agreement or their own value. far need not be filled in.", False),
    ], 1):
        c = ws.cell(i, 1, t); c.font = head if b else Font(); c.alignment = wrap
    ws.column_dimensions["A"].width = 95

    ws = wb.create_sheet(f"Authoring_{len(hand)}")
    cols = ["id", "Question", "Gold (correct)", "Type", "Family", "Why manual",
            "Wikidata candidates (pick no.)", "near (no. or your own)", "checked (2nd)", "note"]
    for j, name in enumerate(cols, 1):
        c = ws.cell(1, j, name); c.font = head; c.alignment = wrap
    for j, w in enumerate([7, 46, 24, 13, 16, 30, 40, 18, 16, 22], 1):
        ws.column_dimensions[chr(64 + j)].width = w
    for ri, (r, reason) in enumerate(hand, 2):
        cands = "\n".join(f"{n}. {c}" for n, c in enumerate(r["_short"], 1)) or "- (write your own)"
        vals = [f'{r["instance_id"]}.{r["hop_idx"]}', r["question"], r["gold_answer"],
                r["gold_kind"], r["family"], REASON_RU.get(reason, reason), cands, "", "", ""]
        for j, v in enumerate(vals, 1):
            c = ws.cell(ri, j, v); c.alignment = wrap
            if j == 8:
                c.fill = hi
    ws.freeze_panes = "A2"

    Path(args.out).parent.mkdir(parents=True, exist_ok=True)
    wb.save(args.out)
    print("->", args.out)


if __name__ == "__main__":
    main()
